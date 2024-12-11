import os
import PyPDF2
import openpyxl
import re

nombre_hoja = "Quote Lines"
columna_inicio = 2  
fila_inicio = 3    
datos = {}

# Define las rutas
carpeta_pdf = 'H:/develop/Sap-scripting/automations/bills/files'
ruta_nuevo_archivo = 'H:/develop/Sap-scripting/automations/bills/format/'
archivo_excel = 'H:/develop/Sap-scripting/automations/bills/format/DefaultTest.xlsx'

def extraer_bloque_codigo(texto):
    # Expresión regular para encontrar el bloque de código deseado
    pattern = re.compile(r'Additional info: Order [\w\-\.\s:]+')
    resultado = pattern.search(texto)
    if resultado:
        return resultado.group(0).replace('Additional info: ', '').strip()
    else:
        return None


def extraer_tabla(texto):
    # Expresión regular para encontrar filas de la tabla
    pattern = re.compile(r'(\d+)\s+([\d,]+\s*H)\s+([\w\s]+?)\s+([\d,]+\s*EUR)\s+([\d.,]+\s*EUR)')
    filas = pattern.findall(texto)
    
    # Limpiar los datos, intercambiar índices y convertir a listas de cadenas
    filas_limpias = []
    for fila in filas:
        item, quantity, description, unit_price, value = fila
        quantity = quantity.replace('H', '').replace(',','.').strip()
        unit_price = unit_price.replace('EUR', '').replace(',','.').strip()
        value = value.replace('EUR', '').strip()
        filas_limpias.append([item, '', description, quantity, unit_price])  # Intercambiar índices 1 y 2 y agregar columna vacía para poder pasar la celda que esta mergeada y no la agarra como tal
    
    return filas_limpias


def leer_pdf(file_path):
    pdf_file = open(file_path, 'rb')
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    texto = ""
    for page in range(len(pdf_reader.pages)):
        texto += pdf_reader.pages[page].extract_text()
    pdf_file.close()
    return texto

def merge_cells(sheet, start_row, start_col, end_row, end_col):
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)


def escribir_tabla_excel(filename, file_path, datos, hoja_nombre, columna_inicio, fila_inicio, aditional_info):


    # Abrir el archivo de Excel existente
    workbook = openpyxl.load_workbook(file_path)
    
    # Seleccionar la hoja especificada
    if hoja_nombre in workbook.sheetnames:
        sheet = workbook[hoja_nombre]
    else:
        raise ValueError(f"La hoja '{hoja_nombre}' no existe en el archivo Excel.")
    
    celda_addiotional_info = sheet.cell(row=22, column=4)

    # Escribir los datos en las celdas especificadas
    for i, fila in enumerate(datos):
        for j, valor in enumerate(fila):
            row = fila_inicio + i
            col = columna_inicio + j
            celda = sheet.cell(row=row, column=col)
            
            # Verificar si la celda está fusionada
            if isinstance(celda, openpyxl.cell.cell.MergedCell):
                # Buscar la celda superior izquierda de la fusión
                for merged_range in sheet.merged_cells.ranges:
                    if (celda.coordinate) in merged_range:
                        top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left_cell.value = valor
                        break 
            else:
                celda.value = valor
    
    celda = sheet.cell(row=fila_inicio, column=columna_inicio)

    if isinstance(celda_addiotional_info, openpyxl.cell.cell.MergedCell):
        # Buscar la celda superior izquierda de la fusión
        for merged_range in sheet.merged_cells.ranges:
            if (celda_addiotional_info.coordinate) in merged_range:
                top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_cell.value = aditional_info
                break 
    else:
        celda_addiotional_info.value = aditional_info

    # Guardar los cambios en el archivo Excel
    nueva_ruta = os.path.join(ruta_nuevo_archivo, filename.replace('.pdf', '.xlsx'))    
    workbook.save(nueva_ruta)

def procesar_tablas_en_carpeta(carpeta_pdf, archivo_excel):
    
    # Recorre todos los archivos en la carpeta
    for filename in os.listdir(carpeta_pdf):
        if filename.endswith(".pdf"):
            file_path = os.path.join(carpeta_pdf, filename)
            texto = leer_pdf(file_path)
            datos_tabla = extraer_tabla(texto)
            aditional_info = extraer_bloque_codigo(texto)
            escribir_tabla_excel(filename, archivo_excel, datos_tabla, nombre_hoja, columna_inicio, fila_inicio, aditional_info)
            print(f"Tabla extraída de '{filename}' y escrita en '{archivo_excel}'.")

# Procesa los archivos PDF y escribe en Excel
procesar_tablas_en_carpeta(carpeta_pdf, archivo_excel)




